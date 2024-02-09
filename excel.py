import openpyxl

wb = openpyxl.load_workbook('videogamesales.xlsx')
ws = wb['vgsales']
print('Total number of rows: '+str(ws.max_row)+'. And total number of columns: '+str(ws.max_column))